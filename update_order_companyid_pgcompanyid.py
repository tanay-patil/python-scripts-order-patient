import requests
import openpyxl
import json
import os
from datetime import datetime

def read_order_ids_from_excel(excel_file):
    """Read order IDs from Excel file"""
    try:
        if not os.path.exists(excel_file):
            print(f"Excel file '{excel_file}' not found!")
            return []
            
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        order_ids = []
        
        # Find the column containing order IDs
        headers = []
        for col in range(1, ws.max_column + 1):
            header_value = ws.cell(row=1, column=col).value
            if header_value:
                headers.append(str(header_value).lower())
        
        order_id_col = None
        for i, header in enumerate(headers):
            if 'order' in header.lower() and 'id' in header.lower():
                order_id_col = i + 1
                break
        
        if order_id_col is None:
            # If no header found, assume first column contains order IDs
            order_id_col = 1
            print("No 'order id' header found, using first column")
        
        # Read order IDs from the identified column
        for row in range(2, ws.max_row + 1):  # Start from row 2 to skip header
            cell_value = ws.cell(row=row, column=order_id_col).value
            if cell_value is not None:
                order_ids.append(str(cell_value).strip())
        
        wb.close()
        print(f"Found {len(order_ids)} order IDs in {excel_file}")
        return order_ids
        
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return []

def fetch_order_by_id(order_id):
    """Fetch order details by order ID"""
    api_url = f"https://dawavorderpatient-hqe2apddbje9gte0.eastus-01.azurewebsites.net/api/Order/{order_id}"
    
    try:
        response = requests.get(api_url)
        response.raise_for_status()
        return response.json()
    except requests.RequestException as e:
        print(f"Error fetching order {order_id}: {e}")
        return None
    except json.JSONDecodeError as e:
        print(f"Error parsing JSON response for order {order_id}: {e}")
        return None

def fetch_patient_details(patient_id):
    """Fetch patient details by patient ID"""
    api_url = f"https://dawavorderpatient-hqe2apddbje9gte0.eastus-01.azurewebsites.net/api/Patient/get-patient/{patient_id}"
    
    try:
        response = requests.get(api_url)
        response.raise_for_status()
        return response.json()
    except requests.RequestException as e:
        print(f"Error fetching patient {patient_id}: {e}")
        return None
    except json.JSONDecodeError as e:
        print(f"Error parsing JSON response for patient {patient_id}: {e}")
        return None

def update_order(order_id, order_data):
    """Update order using PUT API"""
    api_url = f"https://dawavorderpatient-hqe2apddbje9gte0.eastus-01.azurewebsites.net/api/Order/{order_id}"
    
    try:
        # Set headers for PUT request
        headers = {'Content-Type': 'application/json'}
        
        # Debug: Print the data being sent
        print(f"Updating order {order_id} with data keys: {list(order_data.keys())}")
        
        response = requests.put(api_url, json=order_data, headers=headers)
        response.raise_for_status()
        print(f"Successfully updated order {order_id}")
        return True
    except requests.RequestException as e:
        print(f"Error updating order {order_id}: {e}")
        if hasattr(e, 'response') and e.response is not None:
            print(f"Response status: {e.response.status_code}")
            print(f"Response text: {e.response.text}")
        return False

def check_and_update_order_company_ids(order_id):
    """Check order for missing company IDs and update from patient data if needed"""
    
    print(f"\nProcessing order ID: {order_id}")
    
    # Fetch order details
    order = fetch_order_by_id(order_id)
    if not order:
        print(f"Could not fetch order {order_id}")
        return {
            'order_id': order_id,
            'status': 'error',
            'message': 'Could not fetch order details',
            'updated': False
        }
    
    # Check if companyId and pgCompanyId are missing
    company_id = order.get('companyId')
    pg_company_id = order.get('pgCompanyId')
    patient_id = order.get('patientId')
    
    print(f"Order {order_id}: companyId={company_id}, pgCompanyId={pg_company_id}, patientId={patient_id}")
    
    # Check if both company IDs are present
    if company_id and pg_company_id:
        print(f"Order {order_id} already has both company IDs")
        return {
            'order_id': order_id,
            'status': 'no_update_needed',
            'message': 'Both companyId and pgCompanyId already present',
            'updated': False
        }
    
    # Check if patient ID is available
    if not patient_id:
        print(f"Order {order_id} has no patient ID")
        return {
            'order_id': order_id,
            'status': 'error',
            'message': 'No patient ID in order',
            'updated': False
        }
    
    # Fetch patient details
    print(f"Fetching patient details for patient ID: {patient_id}")
    patient = fetch_patient_details(patient_id)
    if not patient:
        print(f"Could not fetch patient {patient_id}")
        return {
            'order_id': order_id,
            'status': 'error',
            'message': f'Could not fetch patient details for {patient_id}',
            'updated': False
        }
    
    # Extract company IDs from patient data
    patient_company_id = None
    patient_pg_company_id = None
    
    if 'agencyInfo' in patient:
        patient_company_id = patient['agencyInfo'].get('companyId')
        patient_pg_company_id = patient['agencyInfo'].get('pgcompanyID')  # Note: different case in patient API
    
    print(f"Patient {patient_id}: companyId={patient_company_id}, pgCompanyId={patient_pg_company_id}")
    
    # Check if patient has the required company IDs
    if not patient_company_id and not patient_pg_company_id:
        print(f"Patient {patient_id} has no company IDs to use for update")
        return {
            'order_id': order_id,
            'status': 'no_update_available',
            'message': 'Patient has no company IDs to update order with',
            'updated': False
        }
    
    # Determine what needs to be updated
    update_needed = False
    updates_to_apply = {}
    
    if not company_id and patient_company_id:
        updates_to_apply['companyId'] = patient_company_id
        update_needed = True
        print(f"Will update companyId to: {patient_company_id}")
    
    if not pg_company_id and patient_pg_company_id:
        updates_to_apply['pgCompanyId'] = patient_pg_company_id
        update_needed = True
        print(f"Will update pgCompanyId to: {patient_pg_company_id}")
    
    if not update_needed:
        print(f"No update needed for order {order_id}")
        return {
            'order_id': order_id,
            'status': 'no_update_needed',
            'message': 'Order already has available company IDs or patient has no new IDs to provide',
            'updated': False
        }
    
    # Create a clean copy of the order data and apply only the specific updates
    # This ensures we don't lose any existing data
    updated_order = order.copy()  # Create a copy to avoid modifying the original
    
    # Apply only the specific updates
    for key, value in updates_to_apply.items():
        updated_order[key] = value
        print(f"Applied update: {key} = {value}")
    
    # Ensure ALL fields from the original order are preserved
    # Based on the API response structure, include all possible fields
    required_fields = [
        'id', 'orderWAVId', 'orderNo', 'orderDate', 'startOfCare', 'episodeStartDate', 
        'episodeEndDate', 'documentID', 'mrn', 'patientName', 'sentToPhysicianDate', 
        'sentToPhysicianStatus', 'signedByPhysicianDate', 'signedByPhysicianStatus',
        'uploadedSignedOrderDate', 'uploadedSignedOrderStatus', 'uploadedSignedPgOrderDate',
        'uploadedSignedPgOrderStatus', 'cpoMinutes', 'orderUrl', 'documentName', 'ehr',
        'account', 'location', 'remarks', 'patientId', 'companyId', 'pgCompanyId',
        'entityType', 'clinicalJustification', 'billingProvider', 'billingProviderNPI',
        'supervisingProvider', 'supervisingProviderNPI', 'bit64Url', 'daOrderType',
        'daUploadSuccess', 'daResponseStatusCode', 'daResponseDetails', 'createdBy',
        'createdOn', 'updatedBy', 'updatedOn'
    ]
    
    # Ensure all fields from the original order are in the updated order
    for field in required_fields:
        if field in order:
            # Only update if the field doesn't already exist in updated_order or if it's None
            if field not in updated_order or updated_order[field] is None:
                updated_order[field] = order[field]
    
    print(f"Final payload will include {len(updated_order)} fields")
    print(f"Fields being sent: {list(updated_order.keys())}")
    
    # Update the order
    success = update_order(order_id, updated_order)
    if success:
        return {
            'order_id': order_id,
            'status': 'success',
            'message': f'Updated with companyId: {patient_company_id}, pgCompanyId: {patient_pg_company_id}',
            'updated': True,
            'new_company_id': patient_company_id,
            'new_pg_company_id': patient_pg_company_id
        }
    else:
        return {
            'order_id': order_id,
            'status': 'update_failed',
            'message': 'Failed to update order via API',
            'updated': False
        }

def save_results_to_excel(results):
    """Save processing results to Excel file"""
    if not results:
        print("No results to save")
        return None
        
    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Order Update Results"
    
    # Write headers
    headers = ['Order ID', 'Status', 'Message', 'Updated', 'New Company ID', 'New PG Company ID']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Write data
    for row, result in enumerate(results, 2):
        ws.cell(row=row, column=1, value=result['order_id'])
        ws.cell(row=row, column=2, value=result['status'])
        ws.cell(row=row, column=3, value=result['message'])
        ws.cell(row=row, column=4, value=result['updated'])
        ws.cell(row=row, column=5, value=result.get('new_company_id', ''))
        ws.cell(row=row, column=6, value=result.get('new_pg_company_id', ''))
    
    # Generate filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"order_update_results_{timestamp}.xlsx"
    
    # Save the workbook
    wb.save(filename)
    return filename

def main():
    print("Order Company ID Update Script")
    print("=" * 40)
    
    # Specify the Excel file containing order IDs
    order_excel_file = "orders.xlsx"  # Change this to your actual file name
    
    # Check if the file exists, if not, suggest alternatives
    if not os.path.exists(order_excel_file):
        print(f"File '{order_excel_file}' not found!")
        print("Please make sure you have an Excel file with order IDs.")
        print("Available Excel files in the current directory:")
        for file in os.listdir("."):
            if file.endswith(".xlsx"):
                print(f"  - {file}")
        
        # Ask user to specify the file
        print("\nPlease update the 'order_excel_file' variable in the script with your actual filename.")
        return
    
    # Read order IDs from Excel file
    order_ids = read_order_ids_from_excel(order_excel_file)
    
    if not order_ids:
        print("No order IDs found in the Excel file")
        return
    
    print(f"\nFound {len(order_ids)} order IDs to process")
    
    # Process each order
    results = []
    total_orders = len(order_ids)
    updated_count = 0
    
    for i, order_id in enumerate(order_ids, 1):
        print(f"\nProgress: {i}/{total_orders}")
        result = check_and_update_order_company_ids(order_id)
        results.append(result)
        
        if result['updated']:
            updated_count += 1
    
    # Save results to Excel
    results_filename = save_results_to_excel(results)
    
    # Print summary
    print(f"\n" + "=" * 50)
    print("PROCESSING SUMMARY")
    print("=" * 50)
    print(f"Total orders processed: {total_orders}")
    print(f"Orders updated: {updated_count}")
    print(f"Orders not updated: {total_orders - updated_count}")
    
    # Count by status
    status_counts = {}
    for result in results:
        status = result['status']
        status_counts[status] = status_counts.get(status, 0) + 1
    
    print(f"\nStatus breakdown:")
    for status, count in status_counts.items():
        print(f"  {status}: {count}")
    
    if results_filename:
        print(f"\nDetailed results saved to: {results_filename}")

if __name__ == "__main__":
    main()
